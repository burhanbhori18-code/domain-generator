"""
Script to create the sample Excel template with campaign-based structure
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Campaign names
CAMPAIGN_NAMES = [
    "Surface SMC+ENT",
    "Accelerate",
    "Security",
    "AI Transformation BDM ENT",
    "AI Transformation BDM SMC",
    "Innovate AI Apps & Agents ENT",
    "Innovate AI Apps & Agents SMC",
    "Migrate & Modernize ENT",
    "Migrate & Modernize SMEC",
    "Unify your Data Platform ENT",
    "Unify your Data Platform SMEC",
    "Data Security ENT",
    "Data Security SMC",
    "Modern SecOps ENT",
    "Modern SecOps SMC",
    "Protect Cloud ENT",
    "Protect Cloud SMC",
    "Copilot ENT",
    "Copilot SMC",
    "ERP Transformation ENT",
    "ERP Transformation SMC",
    "Low Code ENT",
    "Low Code SMC",
    "Sales Transformation ENT",
    "Sales Transformation SMC",
    "Scale ENT",
    "Secure AI Productivity ENT",
    "Secure AI Productivity SMC",
    "Service Transformation SMC",
    "Healthcare Copilot Dragon"
]

# Sample countries and domains for campaign sheets
SAMPLE_CAMPAIGN_DATA = {
    "Surface SMC+ENT": [
        ("USA", "microsoft.com"), ("USA", "apple.com"), ("USA", "google.com"),
        ("Germany", "siemens.com"), ("Germany", "bmw.com"),
        ("France", "airbus.com"), ("UK", "bp.com")
    ],
    "Accelerate": [
        ("Singapore", "dbs.com"), ("Singapore", "singtel.com"),
        ("Malaysia", "petronas.com"), ("Malaysia", "maybank.com"),
        ("India", "tcs.com"), ("India", "infosys.com")
    ],
    "Security": [
        ("USA", "cisco.com"), ("USA", "paloaltonetworks.com"),
        ("Israel", "checkpoint.com"), ("UK", "bae.com"),
        ("Japan", "ntt.com")
    ],
    "AI Transformation BDM ENT": [
        ("USA", "nvidia.com"), ("USA", "openai.com"),
        ("Canada", "元shopify.com"), ("Canada", "blackberry.com"),
        ("China", "alibaba.com")
    ],
    # Default data for remaining campaigns
    "default": [
        ("USA", "example.com"), ("USA", "sample.com"),
        ("Germany", "test.com"), ("France", "demo.com"),
        ("UK", "company.com"), ("Singapore", "business.com")
    ]
}

# Create workbook
wb = Workbook()
ws1 = wb.active
ws1.title = 'Sheet1'

# Sheet1 - TAL Entries with new Campaign column
headers1 = ['TAL Name', 'Countries', 'Campaign (Sales Play)', 'File Created', 'Domain Count', 'Duplicate Domain Found']
ws1.append(headers1)

# Style headers
for cell in ws1[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='001B47', end_color='001B47', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Add sample data with various campaigns
sample_tals = [
    ('Tech_APAC', 'Malaysia, Singapore', 'Accelerate'),
    ('Finance_EU', 'Germany, France', 'Surface SMC+ENT'),
    ('Security_NA', 'USA, UK', 'Security'),
    ('AI_Global', 'USA, Canada, China', 'AI Transformation BDM ENT'),
    ('Cloud_EMEA', 'Germany, UK, France', 'Migrate & Modernize ENT'),
]

for tal_name, countries, campaign in sample_tals:
    ws1.append([tal_name, countries, campaign, '', '', ''])

# Set column widths
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 35
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 25

# Create campaign sheets
for campaign_name in CAMPAIGN_NAMES:
    # Create new sheet
    ws_campaign = wb.create_sheet(campaign_name)
    
    # Add headers
    headers_campaign = ['Country', 'Domain']
    ws_campaign.append(headers_campaign)
    
    # Style headers
    for cell in ws_campaign[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='001B47', end_color='001B47', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Add sample data
    if campaign_name in SAMPLE_CAMPAIGN_DATA:
        data = SAMPLE_CAMPAIGN_DATA[campaign_name]
    else:
        data = SAMPLE_CAMPAIGN_DATA["default"]
    
    for country, domain in data:
        ws_campaign.append([country, domain])
    
    # Set column widths
    ws_campaign.column_dimensions['A'].width = 20
    ws_campaign.column_dimensions['B'].width = 30

# Save the template
wb.save('template/sample_template.xlsx')
print('✅ Template created successfully with 30 campaign sheets!')
print(f'   - Sheet1 with Campaign column')
print(f'   - {len(CAMPAIGN_NAMES)} campaign sheets created')
print('   - Sample data added to all sheets')
