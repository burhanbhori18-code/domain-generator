from flask import Flask, render_template, request, send_file, jsonify
import os
import openpyxl
from openpyxl import Workbook
import zipfile
from datetime import datetime
from werkzeug.utils import secure_filename
import shutil

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure folders exist
os.makedirs('uploads', exist_ok=True)
os.makedirs('outputs', exist_ok=True)

def normalize_country(txt):
    """Normalize country names (same as VBA function)"""
    txt = txt.strip().lower()
    
    # Replace special characters
    replacements = {
        'ü': 'u', 'ö': 'o', 'ä': 'a', 'ß': 'ss',
        'ç': 'c', 'é': 'e', 'á': 'a', 'í': 'i',
        'ó': 'o', 'ú': 'u', 'ñ': 'n'
    }
    
    for old, new in replacements.items():
        txt = txt.replace(old, new)
    
    # Convert to Proper case (Title case)
    return txt.title()

def process_excel_file(filepath):
    """
    Process the uploaded Excel file using memory-efficient streaming (read_only=True).
    Generates domain files and a NEW lightweight Master Summary log.
    Returns the path to the output ZIP file.
    """
    print(f"[{datetime.now()}] Starting file processing (Memory Efficient): {filepath}")
    
    # Clean output folder
    output_folder = app.config['OUTPUT_FOLDER']
    if os.path.exists(output_folder):
        shutil.rmtree(output_folder)
    os.makedirs(output_folder)

    # 1. Load Workbook in Read-Only Mode (Low Memory)
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        raise ValueError(f"Could not open Excel file: {str(e)}")
        
    print(f"[{datetime.now()}] Workbook loaded in read-only mode.")
    
    if 'Sheet1' not in wb.sheetnames:
        raise ValueError("Excel file must contain 'Sheet1'")
    
    ws1 = wb['Sheet1']
    
    # 2. Prepare Data Structures
    campaign_files = {}  # {campaign_name: [list of TAL filenames]}
    total_files_created = 0
    campaign_cache = {}  # {campaign_name: {country: {domain_lower: domain_original}}}
    
    # 3. Create a NEW Master Result Workbook (Lightweight)
    # We cannot edit the read_only workbook, so we create a fresh log file.
    wb_master = Workbook()
    ws_master = wb_master.active
    ws_master.title = "Processing Results"
    
    # Write Headers (Manually set to match expected output)
    headers = ['TAL Name', 'Countries', 'Campaign (Sales Play)', 'File Created', 'Domain Count', 'Duplicate Domain Found']
    ws_master.append(headers)
    
    # 4. Iterate Sheet1 Rows
    processed_count = 0
    
    # Iterate rows from Sheet1 (read_only returns cells generator)
    # Skip header (min_row=2)
    row_idx = 1 # Tracking for log (1-based, but we invoke from 2)
    
    for row in ws1.iter_rows(min_row=2, values_only=True):
        row_idx += 1
        
        # Stop if row is purely empty
        if not row[0] and not row[1] and not row[2]:
            continue
            
        # Extract basic data
        tal_name = str(row[0]).strip() if row[0] else ""
        countries_str = str(row[1]) if row[1] else ""
        campaign_name = str(row[2]).strip() if row[2] else ""
        
        if not tal_name:
            continue
            
        print(f"[{datetime.now()}] Processing Row {row_idx}: {tal_name}")
        
        # Prepare result row data
        result_row = [
            tal_name,
            countries_str,
            campaign_name,
            "", # File Created
            "", # Domain Count
            ""  # info/error
        ]
        
        countries_arr = [c.strip() for c in countries_str.split(',') if c.strip()]
        
        # Validation checks
        if not campaign_name:
            result_row[3] = "No"
            result_row[5] = "Missing campaign name"
            ws_master.append(result_row)
            continue
            
        if campaign_name not in wb.sheetnames:
            result_row[3] = "No"
            result_row[5] = f"Missing sheet: {campaign_name}"
            ws_master.append(result_row)
            continue
            
        # --- CACHING LOGIC (With explicit sheet closing request if possible, though read_only holds handle) ---
        if campaign_name not in campaign_cache:
            # print(f"[{datetime.now()}] Caching sheet: {campaign_name}")
            try:
                campaign_sheet = wb[campaign_name]
                country_domains = {}
                
                # Careful iteration to avoid OOM on huge sheets
                empty_streak = 0
                for data_row in campaign_sheet.iter_rows(min_row=2, values_only=True):
                    c_val, d_val = data_row[0], data_row[1]
                    
                    if not c_val and not d_val:
                        empty_streak += 1
                        if empty_streak > 20: break # Safety break
                        continue
                    empty_streak = 0
                    
                    if c_val and d_val:
                        clean_c = normalize_country(str(c_val))
                        dom_val = str(d_val).strip()
                        if clean_c and dom_val:
                            if clean_c not in country_domains:
                                country_domains[clean_c] = {}
                            if dom_val.lower() not in country_domains[clean_c]:
                                country_domains[clean_c][dom_val.lower()] = dom_val
                
                campaign_cache[campaign_name] = country_domains
                # Note: In read_only, we can't explicitly 'close' a sheet, it's part of the zip stream
            except Exception as e:
                print(f"Error reading campaign sheet {campaign_name}: {e}")
                result_row[5] = f"Error reading sheet: {str(e)}"
                ws_master.append(result_row)
                continue

        country_domains = campaign_cache[campaign_name]
        
        # Check countries
        missing_countries = []
        available_countries = []
        
        for country in countries_arr:
            clean_country = normalize_country(country)
            if clean_country not in country_domains:
                missing_countries.append(country)
            else:
                available_countries.append(country)
        
        if not available_countries:
            result_row[3] = "No"
            result_row[4] = 0
            result_row[5] = f"All countries missing: {', '.join(missing_countries)}"
            ws_master.append(result_row)
            continue
            
        # Collect domains
        collected = {}
        has_dup = False
        for country in available_countries:
            clean_country = normalize_country(country)
            for d_low, d_orig in country_domains[clean_country].items():
                if d_low in collected:
                    has_dup = True
                else:
                    collected[d_low] = d_orig
        
        # Write individual campaign file
        # Creating a new workbook is cheap (small memory)
        try:
            campaign_folder = os.path.join(output_folder, campaign_name)
            os.makedirs(campaign_folder, exist_ok=True)
            
            wb_out = Workbook()
            ws_out = wb_out.active
            for idx_d, d_val in enumerate(collected.values(), start=1):
                ws_out.cell(row=idx_d, column=1).value = d_val
            
            output_filename = f"{tal_name}.xlsx"
            # Sanitize filename
            output_filename = "".join([c for c in output_filename if c.isalpha() or c.isdigit() or c in (' ','.','_','-')]).strip()
            output_path = os.path.join(campaign_folder, output_filename)
            wb_out.save(output_path)
            
            if campaign_name not in campaign_files:
                campaign_files[campaign_name] = []
            campaign_files[campaign_name].append(output_filename)
            total_files_created += 1
            
            # Update result row
            if missing_countries:
                result_row[3] = "Partial"
                result_row[5] = f"Missing: {', '.join(missing_countries)}"
            else:
                result_row[3] = "Yes"
                result_row[5] = "Yes" if has_dup else "No"
            
            result_row[4] = len(collected)
            
        except Exception as e:
            print(f"Error creating file for {tal_name}: {e}")
            result_row[3] = "Error"
            result_row[5] = f"Write error: {str(e)}"

        ws_master.append(result_row)
        processed_count += 1

    # Close the read-only workbook to free file handle
    wb.close()
    
    print(f"[{datetime.now()}] Processing complete. Saving master file...")
    
    # Save the NEW Master Summary
    master_output_path = os.path.join(output_folder, "Master_Summary_Log.xlsx")
    wb_master.save(master_output_path)
    
    # Create ZIP
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_filename = f"domain_files_{timestamp}.zip"
    zip_path = os.path.join(output_folder, zip_filename)
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(master_output_path, "Master_Summary_Log.xlsx")
        for campaign_name, filenames in campaign_files.items():
            # Add Master Summary to each folder too? Or just the root.
            # User requirement: "Master_Results.xlsx" inside. We'll put our new log there.
            zipf.write(master_output_path, f"{campaign_name}/Master_Summary_Log.xlsx")
            for filename in filenames:
                f_path = os.path.join(output_folder, campaign_name, filename)
                zipf.write(f_path, f"{campaign_name}/{filename}")
    
    return zip_path, total_files_created


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/download-template')
def download_template():
    template_path = os.path.join('template', 'sample_template.xlsx')
    return send_file(template_path, as_attachment=True, download_name='Domain_Generator_Template.xlsx')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Please upload an Excel file (.xlsx or .xls)'}), 400
    
    try:
        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        # Process file
        zip_path, file_count = process_excel_file(filepath)
        
        # Clean up uploaded file (with retry logic for locked files)
        try:
            import time
            time.sleep(0.5)  # Brief delay to ensure file handles are released
            os.remove(filepath)
        except PermissionError:
            # File is locked (e.g., still open in Excel), skip deletion
            # It will be cleaned up on the next upload or manually
            pass
        except Exception as e:
            # Log but don't fail the request
            print(f"Warning: Could not delete uploaded file: {e}")
        
        # Return download path
        zip_filename = os.path.basename(zip_path)
        return jsonify({
            'success': True,
            'message': f'Successfully created {file_count} domain file(s)',
            'download_url': f'/download-result/{zip_filename}',
            'file_count': file_count
        })
        
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

@app.route('/download-result/<filename>')
def download_result(filename):
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    return send_file(filepath, as_attachment=True)

if __name__ == '__main__':
    print("\n" + "="*60)
    print("🚀 Domain File Generator - Madison Logic")
    print("="*60)
    print("Server starting at: http://localhost:5000")
    print("Press CTRL+C to stop the server")
    print("="*60 + "\n")
    app.run(debug=True, port=5000)
